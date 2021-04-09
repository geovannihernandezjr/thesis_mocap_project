"""
@author GEovanni Hernandez
Created to write and save to excel file the results ran in an experiment.
"""

from openpyxl import load_workbook
import os
from pandas import ExcelWriter
def save_model_summary(file_name):
    def print_to_file(line):
        with open(file_name, 'w+') as f:
            print(line, f)



def save_to_excel(excel_name='', stateful=False, excel_sheet = dict, excel_sheet_name='Default'):
    '''
    The data is saved to an excel
    :param excel_name: string name of excel
    :param stateful: determine if the model is stateful or not
    :param excel_sheet: dictonary containing the data that will be written to excel sheet
    :param excel_sheet_name: name of excel sheet
    :return: NA
    '''
    # Create Pandas Excel writer using Openpyxl as the engine
    writer = ExcelWriter(excel_name, engine='openpyxl')
    if not os.path.exists(excel_name):
        print("File DOES NOT EXIST SO CREATING EXCEL")
        # STATEFUL WITH RESET
        if stateful:

            # Convert the dataframes for each sheet title to an openpyxl Excel object
            # excel_sheet_1['StatefulWithReset'].to_excel(writer, sheet_name='StatefulWithReset', index=False)
            # excel_sheet['Stateful'].to_excel(writer, sheet_name='Stateful', index=False)
            excel_sheet[excel_sheet_name].to_excel(writer, sheet_name=excel_sheet_name, index=False)

        # STATELESS
        elif False is stateful:

            # Convert the dataframes for each sheet title to an openpyxl Excel object
            # excel_sheet_3['Stateless'].to_excel(writer, sheet_name='Stateless', index=False)
            # excel_sheet['Stateless'].to_excel(writer, sheet_name='Stateless', index=False)
            excel_sheet[excel_sheet_name].to_excel(writer, sheet_name=excel_sheet_name, index=False)
        else:
            print(f'Excel Sheet Name is not Stateless or Stateful {excel_sheet_name}')
            excel_sheet[excel_sheet_name].to_excel(writer, sheet_name=excel_sheet_name, index=False)
        # Close the Pandas Excel Writer and output the Excel file
        writer.save()
    else:
        print("FILE EXISTS, OPENING FILE AND ADDING TO IT")
        # try to open an existing workbook
        writer.book = load_workbook(excel_name)
        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        # read existing
        # STATEFUL WITH RESET
        if stateful:

            # if not 'Stateful' in writer.sheets: # if sheet with name of 'Stateful' does not exist
            if not excel_sheet_name in writer.sheets:
                # excel_sheet_1['StatefulWithReset'].to_excel(writer, sheet_name='StatefulWithReset', index=False)
                # excel_sheet['Stateful'].to_excel(writer, sheet_name='Stateful', index=False)
                excel_sheet[excel_sheet_name].to_excel(writer, sheet_name=excel_sheet_name, index=False)


            else:
                # lastrow = writer.sheets['Stateful'].max_row
                lastrow = writer.sheets[excel_sheet_name].max_row

                # Convert the dataframes for each sheet title to an openpyxl Excel object
                # excel_sheet_1['StatefulWithReset'].to_excel(writer, sheet_name='StatefulWithReset',
                #                                             startrow=lastrow, index=False, header=False)
                # excel_sheet['Stateful'].to_excel(writer, sheet_name='Stateful',
                #                                             startrow=lastrow, index=False, header=False)
                excel_sheet[excel_sheet_name].to_excel(writer, sheet_name=excel_sheet_name,
                                                               startrow=lastrow, index=False, header=False)

            # excel_sheet_1.clear()
            # excel_sheet.clear()

        # STATELESS
        elif False is stateful:

            # if not 'Stateless' in writer.sheets:
            if not excel_sheet_name in writer.sheets:
                # excel_sheet_3['Stateless'].to_excel(writer, sheet_name='Stateless', index=False)
                # excel_sheet['Stateless'].to_excel(writer, sheet_name='Stateless', index=False)
                excel_sheet[excel_sheet_name].to_excel(writer, sheet_name=excel_sheet_name, index=False)

            else:
                # lastrow = writer.sheets['Stateless'].max_row
                lastrow = writer.sheets[excel_sheet_name].max_row

                # Convert the dataframes for each sheet title to an openpyxl Excel object
                # excel_sheet_3['Stateless'].to_excel(writer, sheet_name='Stateless', startrow=lastrow, index=False,
                #                                     header=False)
                # excel_sheet['Stateless'].to_excel(writer, sheet_name='Stateless', startrow=lastrow, index=False,
                #                                     header=False)
                excel_sheet[excel_sheet_name].to_excel(writer, sheet_name=excel_sheet_name,
                                                       startrow=lastrow, index=False, header=False)

            # excel_sheet_3.clear()
            # excel_sheet.clear()

        writer.save()
        writer.close()
